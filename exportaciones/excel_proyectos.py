"""
Exportación a Excel del inventario de proyectos (vigencia única o consolidado).

Formatea el archivo con la paleta corporativa: masthead azul, eyebrow naranja,
filas alternas, formatos de número y porcentaje, columnas con anchos
predefinidos para las columnas conocidas y autoajuste para el resto.
"""
import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =========================================================================
# Paleta en formato openpyxl (sin '#')
# =========================================================================
XL_BLUE_DARK = "003D6C"
XL_ORANGE    = "CF7000"
XL_INK       = "0D1B2A"
XL_INK_MUTE  = "4A5A6A"
XL_HAIRLINE  = "E3E3E1"
XL_PAPER     = "FFFFFF"
XL_ALT_ROW   = "F6F6F5"


def generar_excel_proyectos(
    df_proyectos: pd.DataFrame,
    titulo: str,
    subtitulo: str,
) -> bytes:
    """Genera un xlsx formateado corporativamente con la tabla de proyectos.

    ``df_proyectos`` debe traer las columnas que ya construimos en la pestaña +
    opcionalmente 'Vigencia PI' al inicio (cuando es consolidado de todas
    las vigencias).
    """
    thin = Border(
        left=Side(style="thin", color=XL_HAIRLINE),
        right=Side(style="thin", color=XL_HAIRLINE),
        top=Side(style="thin", color=XL_HAIRLINE),
        bottom=Side(style="thin", color=XL_HAIRLINE),
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Proyectos"
    ws.sheet_view.showGridLines = False

    n_cols = len(df_proyectos.columns)

    # --- Masthead ---
    ws.cell(row=1, column=1, value=titulo)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(row=1, column=1)
    c.font = Font(name="Montserrat", bold=True, size=18, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=XL_BLUE_DARK)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 38

    ws.cell(row=2, column=1, value=subtitulo)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    c = ws.cell(row=2, column=1)
    c.font = Font(name="Open Sans", italic=True, size=10, color=XL_INK_MUTE)
    c.fill = PatternFill("solid", fgColor=XL_PAPER)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 22

    # Línea naranja decorativa
    for col in range(1, n_cols + 1):
        c = ws.cell(row=3, column=col)
        c.fill = PatternFill("solid", fgColor=XL_ORANGE)
    ws.row_dimensions[3].height = 4

    # --- Encabezado tabla ---
    header_row = 5
    for c_idx, col_name in enumerate(df_proyectos.columns, start=1):
        cell = ws.cell(row=header_row, column=c_idx, value=str(col_name))
        cell.font = Font(name="Montserrat", bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=XL_BLUE_DARK)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style="thin", color=XL_BLUE_DARK),
            right=Side(style="thin", color=XL_BLUE_DARK),
            top=Side(style="thin", color=XL_BLUE_DARK),
            bottom=Side(style="medium", color=XL_ORANGE),
        )
    ws.row_dimensions[header_row].height = 32

    # --- Cuerpo ---
    pct_cols = {"Avance"}
    num_cols = {"Meta", "Ejecutado"}
    for r_idx, (_, row) in enumerate(df_proyectos.iterrows(), start=header_row + 1):
        alt = (r_idx - header_row) % 2 == 0
        for c_idx, col_name in enumerate(df_proyectos.columns, start=1):
            v = row[col_name]
            if pd.isna(v):
                v = None
            cell = ws.cell(row=r_idx, column=c_idx, value=v)
            cell.font = Font(name="Open Sans", size=10, color=XL_INK)
            if alt:
                cell.fill = PatternFill("solid", fgColor=XL_ALT_ROW)
            cell.border = thin
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if col_name in pct_cols:
                cell.number_format = "0.00%"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_name in num_cols:
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")

    # --- Anchos ---
    anchos = {
        "Vigencia PI": 12, "Codigo Meta": 14, "Línea Estratégica": 28,
        "Sector PDD": 24, "Programa PDD": 32,
        "Nombre del Proyecto": 50, "BPIN": 18, "Indicador": 42,
        "Tipo de Banco": 22, "Meta": 14, "Ejecutado": 14, "Avance": 12,
    }
    for c_idx, col_name in enumerate(df_proyectos.columns, start=1):
        letra = get_column_letter(c_idx)
        if col_name in anchos:
            ws.column_dimensions[letra].width = anchos[col_name]
        else:
            serie = df_proyectos[col_name].astype(str)
            max_len = max(len(str(col_name)), serie.str.len().max() if not serie.empty else 0)
            ws.column_dimensions[letra].width = min(max(12, max_len + 2), 45)

    # Congelar paneles bajo el encabezado
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
