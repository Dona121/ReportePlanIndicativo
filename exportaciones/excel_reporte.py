"""
Exportación a Excel del reporte completo del tablero, formateado con
la identidad visual corporativa.

El archivo incluye ocho hojas: Portada, Financiera por Fuente,
Financiera Cuatrienio, Por Categoría PDD, Ejecución Física,
Por Dependencia, Detalle por Meta y Proyectos.
"""
import io

import pandas as pd
import polars as pl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from procesamiento.proyectos import construir_proyectos


def generar_reporte_excel(
    datos: dict,
    vigencia: str,
    ejec_financ_tipo,
    ejec_acumulada_tipo,
    categorias_pdd,
    ejec_dependencia,
    avances_fisicos,
) -> bytes:
    """Genera un Excel con todas las hojas del reporte, formateado con la paleta corporativa."""
    # ---- Paleta en formato openpyxl (sin '#') ----
    XL_BLUE_DARK  = "003D6C"
    XL_BLUE       = "1754AB"
    XL_ORANGE     = "CF7000"
    XL_PAPER      = "FFFFFF"
    XL_BEIGE      = "EDEDEB"
    XL_HAIRLINE   = "E3E3E1"
    XL_INK        = "0D1B2A"
    XL_INK_MUTE   = "4A5A6A"

    # ---- Estilos base ----
    thin_border = Border(
        left=Side(style="thin", color=XL_HAIRLINE),
        right=Side(style="thin", color=XL_HAIRLINE),
        top=Side(style="thin", color=XL_HAIRLINE),
        bottom=Side(style="thin", color=XL_HAIRLINE),
    )

    def estilo_header(cell):
        cell.font = Font(name="Montserrat", bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=XL_BLUE_DARK)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style="thin", color=XL_BLUE_DARK),
            right=Side(style="thin", color=XL_BLUE_DARK),
            top=Side(style="thin", color=XL_BLUE_DARK),
            bottom=Side(style="medium", color=XL_ORANGE),
        )

    def estilo_dato(cell, alt=False):
        cell.font = Font(name="Open Sans", size=10, color=XL_INK)
        if alt:
            cell.fill = PatternFill("solid", fgColor="F6F6F5")
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    def estilo_total(cell):
        cell.font = Font(name="Montserrat", bold=True, color=XL_BLUE_DARK, size=10)
        cell.fill = PatternFill("solid", fgColor=XL_BEIGE)
        cell.border = Border(
            left=Side(style="thin", color=XL_HAIRLINE),
            right=Side(style="thin", color=XL_HAIRLINE),
            top=Side(style="medium", color=XL_BLUE_DARK),
            bottom=Side(style="thin", color=XL_HAIRLINE),
        )
        cell.alignment = Alignment(vertical="center")

    def escribir_tabla(ws, df: pd.DataFrame, start_row: int, fila_total: dict = None,
                       columnas_pct=None, columnas_money=None, columnas_num=None):
        """Escribe un DataFrame con estilos institucionales desde start_row."""
        columnas_pct = columnas_pct or []
        columnas_money = columnas_money or []
        columnas_num = columnas_num or []

        # Encabezados
        for c_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=start_row, column=c_idx, value=str(col_name))
            estilo_header(cell)
        ws.row_dimensions[start_row].height = 32

        # Filas de datos
        for r_idx, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
            alt = (r_idx - start_row) % 2 == 0
            for c_idx, col_name in enumerate(df.columns, start=1):
                v = row[col_name]
                if pd.isna(v):
                    v = None
                cell = ws.cell(row=r_idx, column=c_idx, value=v)
                estilo_dato(cell, alt=alt)
                if col_name in columnas_pct:
                    cell.number_format = "0.00%"
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif col_name in columnas_money:
                    cell.number_format = '"$ "#,##0'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif col_name in columnas_num:
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right", vertical="center")

        last_row = start_row + len(df)

        # Fila de total
        if fila_total:
            last_row += 1
            for c_idx, col_name in enumerate(df.columns, start=1):
                v = fila_total.get(col_name, "")
                if c_idx == 1 and not v:
                    v = "TOTAL"
                cell = ws.cell(row=last_row, column=c_idx, value=v)
                estilo_total(cell)
                if col_name in columnas_pct:
                    cell.number_format = "0.00%"
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif col_name in columnas_money:
                    cell.number_format = '"$ "#,##0'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif col_name in columnas_num:
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right", vertical="center")

        return last_row

    def ajustar_anchos(ws, df: pd.DataFrame, anchos_especificos=None):
        anchos_especificos = anchos_especificos or {}
        for c_idx, col_name in enumerate(df.columns, start=1):
            letra = get_column_letter(c_idx)
            if col_name in anchos_especificos:
                ws.column_dimensions[letra].width = anchos_especificos[col_name]
            else:
                serie = df[col_name].astype(str)
                max_len = max(len(str(col_name)), serie.str.len().max() if not serie.empty else 0)
                ws.column_dimensions[letra].width = min(max(12, max_len + 2), 45)

    def agregar_titulo(ws, titulo: str, subtitulo: str, start_row: int = 1, span: int = 6):
        # Franja superior azul oscura con el título
        ws.cell(row=start_row, column=1, value=titulo)
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=span)
        tcell = ws.cell(row=start_row, column=1)
        tcell.font = Font(name="Montserrat", bold=True, size=16, color="FFFFFF")
        tcell.fill = PatternFill("solid", fgColor=XL_BLUE_DARK)
        tcell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[start_row].height = 36

        # Franja de eyebrow naranja
        ws.cell(row=start_row + 1, column=1, value=subtitulo)
        ws.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 1, end_column=span)
        scell = ws.cell(row=start_row + 1, column=1)
        scell.font = Font(name="Open Sans", italic=True, size=10, color=XL_INK_MUTE)
        scell.fill = PatternFill("solid", fgColor=XL_PAPER)
        scell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[start_row + 1].height = 22

        # Franja naranja divisoria de 2 px equivalentes
        ws.cell(row=start_row + 2, column=1, value="")
        ws.merge_cells(start_row=start_row + 2, start_column=1, end_row=start_row + 2, end_column=span)
        dcell = ws.cell(row=start_row + 2, column=1)
        dcell.fill = PatternFill("solid", fgColor=XL_ORANGE)
        ws.row_dimensions[start_row + 2].height = 4

        return start_row + 4  # fila siguiente útil

    # ---- Crear workbook ----
    wb = Workbook()

    # =======================================================
    # Hoja 1: Portada
    # =======================================================
    ws = wb.active
    ws.title = "Portada"
    ws.sheet_view.showGridLines = False

    # Cabecera estilo masthead
    ws.cell(row=2, column=2, value="INFORME DE SEGUIMIENTO")
    ws.merge_cells("B2:G2")
    c = ws.cell(row=2, column=2)
    c.font = Font(name="Montserrat", bold=True, size=9, color=XL_ORANGE)
    c.alignment = Alignment(horizontal="left")

    ws.cell(row=4, column=2, value="Plan Indicativo 2024—2027")
    ws.merge_cells("B4:G4")
    c = ws.cell(row=4, column=2)
    c.font = Font(name="Montserrat", bold=True, size=28, color=XL_BLUE_DARK)
    c.alignment = Alignment(horizontal="left")
    ws.row_dimensions[4].height = 42

    ws.cell(row=5, column=2, value=f"Vigencia en análisis: {vigencia}")
    ws.merge_cells("B5:G5")
    c = ws.cell(row=5, column=2)
    c.font = Font(name="Open Sans", italic=True, size=12, color=XL_INK_MUTE)
    c.alignment = Alignment(horizontal="left")

    # Línea naranja decorativa
    for col in range(2, 8):
        c = ws.cell(row=7, column=col)
        c.fill = PatternFill("solid", fgColor=XL_ORANGE)
    ws.row_dimensions[7].height = 4

    # Bloque de descripción
    ws.cell(row=9, column=2,
            value="Consolida la programación y ejecución de los indicadores de producto "
                  "y sus fuentes de financiación. Los datos de 2024 y 2025 corresponden "
                  "a vigencias cerradas; los archivos de 2026 se actualizan en el "
                  "repositorio del sistema.")
    ws.merge_cells("B9:G11")
    c = ws.cell(row=9, column=2)
    c.font = Font(name="Open Sans", size=11, color=XL_INK)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Tarjeta KPI
    kpis = [
        (f"Programación {vigencia}",
         ejec_financ_tipo.select(pl.col(f"Programación Financiera {vigencia}").sum()).item() or 0,
         "money"),
        (f"Ejecución {vigencia}",
         ejec_financ_tipo.select(pl.col(f"Ejecución Financiera {vigencia}").sum()).item() or 0,
         "money"),
        ("Programación Cuatrienio",
         ejec_acumulada_tipo.select(pl.col("Programación Cuatrienio").sum()).item() or 0,
         "money"),
        ("Ejecución Acumulada",
         ejec_acumulada_tipo.select(pl.col("Ejecución Financiera Acumulada").sum()).item() or 0,
         "money"),
        (f"Avance ponderado {vigencia}",
         avances_fisicos["avance_vig_ponderado"] or 0, "pct"),
        ("Avance ponderado cuatrienio",
         avances_fisicos["avance_cuatrienio_total"] or 0, "pct"),
    ]

    fila_kpi = 14
    for i, (label, valor, tipo) in enumerate(kpis):
        col = 2 + (i % 3) * 2
        row = fila_kpi + (i // 3) * 4

        # Etiqueta
        ws.cell(row=row, column=col, value=label.upper())
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        c = ws.cell(row=row, column=col)
        c.font = Font(name="Montserrat", bold=True, size=8, color=XL_INK_MUTE)
        c.fill = PatternFill("solid", fgColor="FFFFFF")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = Border(
            left=Side(style="medium", color=XL_BLUE),
            top=Side(style="thin", color=XL_HAIRLINE),
            right=Side(style="thin", color=XL_HAIRLINE),
        )

        # Valor
        ws.cell(row=row + 1, column=col, value=valor)
        ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 2, end_column=col + 1)
        c = ws.cell(row=row + 1, column=col)
        c.font = Font(name="Montserrat", bold=True, size=18, color=XL_INK)
        c.fill = PatternFill("solid", fgColor="FFFFFF")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = Border(
            left=Side(style="medium", color=XL_BLUE),
            bottom=Side(style="thin", color=XL_HAIRLINE),
            right=Side(style="thin", color=XL_HAIRLINE),
        )
        if tipo == "money":
            c.number_format = '"$ "#,##0'
        elif tipo == "pct":
            c.number_format = "0.00%"

    # Anchos
    for col_letra in ["A", "B", "C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col_letra].width = 18
    ws.column_dimensions["A"].width = 3

    # =======================================================
    # Hoja 2: Ejecución Financiera por Fuente — Vigencia
    # =======================================================
    ws = wb.create_sheet("Financiera por Fuente")
    ws.sheet_view.showGridLines = False

    df_fin_vig = ejec_financ_tipo.to_pandas()[
        ["Tipo Fuente", "Clasificación Recursos",
         f"Programación Financiera {vigencia}",
         f"Ejecución Financiera {vigencia}",
         "Porcentaje de Ejecución Financiera"]
    ].rename(columns={
        f"Programación Financiera {vigencia}": f"Programación {vigencia}",
        f"Ejecución Financiera {vigencia}": f"Ejecución {vigencia}",
        "Porcentaje de Ejecución Financiera": "% Ejecución",
    })

    inicio = agregar_titulo(
        ws,
        f"Ejecución financiera — Vigencia {vigencia}",
        "Programación vs ejecución por clasificación de recursos",
        span=len(df_fin_vig.columns),
    )

    total_prog = df_fin_vig[f"Programación {vigencia}"].sum()
    total_ejec = df_fin_vig[f"Ejecución {vigencia}"].sum()
    fila_tot = {
        f"Programación {vigencia}": total_prog,
        f"Ejecución {vigencia}": total_ejec,
        "% Ejecución": (total_ejec / total_prog) if total_prog else 0,
    }
    escribir_tabla(
        ws, df_fin_vig, inicio, fila_total=fila_tot,
        columnas_money=[f"Programación {vigencia}", f"Ejecución {vigencia}"],
        columnas_pct=["% Ejecución"],
    )
    ajustar_anchos(ws, df_fin_vig, anchos_especificos={
        "Tipo Fuente": 32, "Clasificación Recursos": 28,
        f"Programación {vigencia}": 22, f"Ejecución {vigencia}": 22, "% Ejecución": 16,
    })

    # =======================================================
    # Hoja 3: Ejecución Financiera Cuatrienio
    # =======================================================
    ws = wb.create_sheet("Financiera Cuatrienio")
    ws.sheet_view.showGridLines = False

    df_acum = ejec_acumulada_tipo.to_pandas().copy()
    df_acum["% Avance Cuatrienio"] = df_acum.apply(
        lambda r: (r["Ejecución Financiera Acumulada"] / r["Programación Cuatrienio"])
        if r["Programación Cuatrienio"] else 0, axis=1
    )
    df_acum = df_acum[[
        "Tipo Fuente", "Clasificación Recursos",
        "Programación Cuatrienio",
        "Ejecución Financiera 2024", "Ejecución Financiera 2025", "Ejecución Financiera 2026",
        "Ejecución Financiera Acumulada", "% Avance Cuatrienio",
    ]]

    inicio = agregar_titulo(
        ws, "Ejecución financiera — Cuatrienio",
        "Ejecución acumulada por fuente desde 2024 hasta 2026",
        span=len(df_acum.columns),
    )

    fila_tot = {
        "Programación Cuatrienio": df_acum["Programación Cuatrienio"].sum(),
        "Ejecución Financiera 2024": df_acum["Ejecución Financiera 2024"].sum(),
        "Ejecución Financiera 2025": df_acum["Ejecución Financiera 2025"].sum(),
        "Ejecución Financiera 2026": df_acum["Ejecución Financiera 2026"].sum(),
        "Ejecución Financiera Acumulada": df_acum["Ejecución Financiera Acumulada"].sum(),
        "% Avance Cuatrienio": (
            df_acum["Ejecución Financiera Acumulada"].sum() / df_acum["Programación Cuatrienio"].sum()
            if df_acum["Programación Cuatrienio"].sum() else 0
        ),
    }
    money_cols = ["Programación Cuatrienio", "Ejecución Financiera 2024",
                  "Ejecución Financiera 2025", "Ejecución Financiera 2026",
                  "Ejecución Financiera Acumulada"]
    escribir_tabla(ws, df_acum, inicio, fila_total=fila_tot,
                   columnas_money=money_cols, columnas_pct=["% Avance Cuatrienio"])
    ajustar_anchos(ws, df_acum, anchos_especificos={"Tipo Fuente": 32, "Clasificación Recursos": 28})

    # =======================================================
    # Hoja 4: Ejecución Financiera por Categorías del PDD
    # =======================================================
    ws = wb.create_sheet("Por Categoría PDD")
    ws.sheet_view.showGridLines = False

    fila = 1
    for cat_key, cat_label, col_grupo in [
        ("lineas",    "Líneas Estratégicas", "Línea Estratégica"),
        ("sectores",  "Sectores PDD",        "Sector PDD"),
        ("programas", "Programas PDD",       "Programa PDD"),
    ]:
        df_c = categorias_pdd[cat_key].to_pandas()[
            [col_grupo, f"Programación Financiera {vigencia}",
             f"Ejecución Financiera {vigencia}", "Porcentaje de Ejecución Financiera"]
        ].rename(columns={
            f"Programación Financiera {vigencia}": f"Programación {vigencia}",
            f"Ejecución Financiera {vigencia}": f"Ejecución {vigencia}",
            "Porcentaje de Ejecución Financiera": "% Ejecución",
        })
        fila = agregar_titulo(
            ws, cat_label, f"Programación vs ejecución financiera — {vigencia}",
            start_row=fila, span=len(df_c.columns),
        )
        tot_prog = df_c[f"Programación {vigencia}"].sum()
        tot_ejec = df_c[f"Ejecución {vigencia}"].sum()
        fila_tot = {
            f"Programación {vigencia}": tot_prog,
            f"Ejecución {vigencia}": tot_ejec,
            "% Ejecución": (tot_ejec / tot_prog) if tot_prog else 0,
        }
        ultima = escribir_tabla(
            ws, df_c, fila, fila_total=fila_tot,
            columnas_money=[f"Programación {vigencia}", f"Ejecución {vigencia}"],
            columnas_pct=["% Ejecución"],
        )
        fila = ultima + 3  # espacio entre tablas

    # Ajuste de anchos (columnas típicas)
    anchos = {"A": 48, "B": 22, "C": 22, "D": 16}
    for letra, w in anchos.items():
        ws.column_dimensions[letra].width = w

    # =======================================================
    # Hoja 5: Ejecución Física por Línea/Sector
    # =======================================================
    ws = wb.create_sheet("Ejecución Física")
    ws.sheet_view.showGridLines = False

    fila = 1
    for grupo, df_src, col_grupo, titulo in [
        ("Líneas Estratégicas — Vigencia",
         avances_fisicos["avance_vig_lineas"].to_pandas(), "Línea Estratégica",
         f"Avance físico por Línea Estratégica — {vigencia}"),
        ("Líneas Estratégicas — Cuatrienio",
         avances_fisicos["avance_cuatri_lineas"].to_pandas(), "Línea Estratégica",
         "Avance físico por Línea Estratégica — Cuatrienio"),
        ("Sectores PDD — Vigencia",
         avances_fisicos["avance_vig_sectores"].to_pandas(), "Sector PDD",
         f"Avance físico por Sector PDD — {vigencia}"),
        ("Sectores PDD — Cuatrienio",
         avances_fisicos["avance_cuatri_sectores"].to_pandas(), "Sector PDD",
         "Avance físico por Sector PDD — Cuatrienio"),
    ]:
        df_g = df_src[[col_grupo, "% Aporte Cumplimiento PDD",
                       "Sobre Numero Total de Indicadores", "% Eficacia Operativa"]].copy()
        df_g = df_g.rename(columns={
            "% Aporte Cumplimiento PDD": "Aporte PDD",
            "Sobre Numero Total de Indicadores": "Peso relativo",
            "% Eficacia Operativa": "Eficacia Operativa",
        })
        fila = agregar_titulo(ws, grupo, titulo, start_row=fila, span=len(df_g.columns))
        ultima = escribir_tabla(
            ws, df_g, fila,
            columnas_pct=["Aporte PDD", "Peso relativo", "Eficacia Operativa"],
        )
        fila = ultima + 3

    ws.column_dimensions["A"].width = 48
    for letra in ["B", "C", "D"]:
        ws.column_dimensions[letra].width = 20

    # =======================================================
    # Hoja 6: Ejecución por Dependencia
    # =======================================================
    ws = wb.create_sheet("Por Dependencia")
    ws.sheet_view.showGridLines = False

    df_dep = ejec_dependencia.to_pandas()[
        ["Varias Secretarías", "Dependencia Responsable",
         f"Metas Programadas {vigencia}", f"Metas Cumplidas al 100% {vigencia}",
         f"Porcentaje de Ejecución {vigencia}", "Porcentaje de Ejecución Acumulada"]
    ].rename(columns={
        f"Metas Programadas {vigencia}": f"Programadas {vigencia}",
        f"Metas Cumplidas al 100% {vigencia}": "Cumplidas 100%",
        f"Porcentaje de Ejecución {vigencia}": f"Avance {vigencia}",
        "Porcentaje de Ejecución Acumulada": "Avance acumulado",
    })

    inicio = agregar_titulo(
        ws, "Ejecución por Dependencia Responsable",
        f"Desempeño físico por secretaría — {vigencia}",
        span=len(df_dep.columns),
    )
    escribir_tabla(
        ws, df_dep, inicio,
        columnas_pct=[f"Avance {vigencia}", "Avance acumulado"],
    )
    ajustar_anchos(ws, df_dep, anchos_especificos={
        "Varias Secretarías": 22, "Dependencia Responsable": 40,
    })

    # =======================================================
    # Hoja 7: Detalle por Meta
    # =======================================================
    ws = wb.create_sheet("Detalle por Meta")
    ws.sheet_view.showGridLines = False

    df_det = datos["prog_fisica_financiera"].to_pandas()[[
        "Codigo Meta", "Línea Estratégica", "Sector PDD", "Programa PDD",
        "Indicador de producto principal", "Responsable",
        "Meta de cuatrienio",
        f"Meta Física Esperada {vigencia}",
        f"EJECUCIÓN {vigencia}",
        f"PORCENTAJE DE EJECUCIÓN {vigencia}",
        "EJECUCIÓN ACUMULADA", "PORCENTAJE DE EJECUCIÓN ACUMULADA",
        "CATEGORÍA DE EJECUCIÓN ACUMULADA",
    ]].rename(columns={
        "Indicador de producto principal": "Indicador",
        f"Meta Física Esperada {vigencia}": f"Meta {vigencia}",
        f"EJECUCIÓN {vigencia}": f"Ejecución {vigencia}",
        f"PORCENTAJE DE EJECUCIÓN {vigencia}": f"Avance {vigencia}",
        "EJECUCIÓN ACUMULADA": "Ejec. acumulada",
        "PORCENTAJE DE EJECUCIÓN ACUMULADA": "Avance acumulado",
        "CATEGORÍA DE EJECUCIÓN ACUMULADA": "Categoría",
    })

    inicio = agregar_titulo(
        ws, "Detalle por Meta",
        f"Inventario de indicadores de producto — Vigencia {vigencia}",
        span=len(df_det.columns),
    )
    escribir_tabla(
        ws, df_det, inicio,
        columnas_pct=[f"Avance {vigencia}", "Avance acumulado"],
        columnas_num=["Meta de cuatrienio", f"Meta {vigencia}",
                      f"Ejecución {vigencia}", "Ejec. acumulada"],
    )
    ws.freeze_panes = ws.cell(row=inicio + 1, column=1)
    ajustar_anchos(ws, df_det, anchos_especificos={
        "Codigo Meta": 14, "Línea Estratégica": 30, "Sector PDD": 26,
        "Programa PDD": 36, "Indicador": 40, "Responsable": 28, "Categoría": 18,
    })

    # =======================================================
    # Hoja 8: Proyectos
    # =======================================================
    df_proy = construir_proyectos(datos, vigencia).to_pandas()
    if not df_proy.empty:
        ws = wb.create_sheet("Proyectos")
        ws.sheet_view.showGridLines = False

        # Construye Indicador como "código — nombre" (igual que en la pestaña)
        def _fmt_ind(row):
            cod = row.get("Código del indicador principal")
            nom = row.get("Indicador de producto principal")
            cod = "" if pd.isna(cod) else str(cod).strip()
            nom = "" if pd.isna(nom) else str(nom).strip()
            if cod and nom:
                return f"{cod} — {nom}"
            return cod or nom or ""

        df_proy_xl = df_proy.copy()
        df_proy_xl["Indicador"] = df_proy_xl.apply(_fmt_ind, axis=1)
        df_proy_xl = df_proy_xl[[
            "Codigo Meta", "Línea Estratégica", "Sector PDD", "Programa PDD",
            "Nombre del Proyecto", "BPIN", "Indicador",
            "Tipo de Banco", "Meta", "Ejecutado",
        ]]
        df_proy_xl["Avance"] = df_proy_xl.apply(
            lambda r: (r["Ejecutado"] / r["Meta"])
            if pd.notna(r["Meta"]) and pd.notna(r["Ejecutado"]) and r["Meta"] != 0
            else None,
            axis=1,
        )

        inicio = agregar_titulo(
            ws, "Proyectos y Gestiones",
            f"Inventario extraído del Plan Indicativo — Vigencia {vigencia}",
            span=len(df_proy_xl.columns),
        )
        escribir_tabla(
            ws, df_proy_xl, inicio,
            columnas_pct=["Avance"],
            columnas_num=["Meta", "Ejecutado"],
        )
        ws.freeze_panes = ws.cell(row=inicio + 1, column=1)
        ajustar_anchos(ws, df_proy_xl, anchos_especificos={
            "Codigo Meta": 14, "Nombre del Proyecto": 50, "Indicador": 42,
            "Tipo de Banco": 22,
            "Línea Estratégica": 28, "Sector PDD": 24, "Programa PDD": 32,
        })

    # ---- Serialización ----
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
